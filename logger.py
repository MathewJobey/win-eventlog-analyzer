#!/usr/bin/env python3
"""
logger.py

Terminal UI to select Windows Event Log and a start/end datetime range,
read events from that log (legacy win32evtlog API), report counts, then
aggregate by normalized EventID and write an Excel report named log_analysis.xlsx.

Compatible with Python 3.7+.
"""

from __future__ import annotations

import sys
import platform
import datetime
import traceback
import os
import shutil
from typing import Tuple, List, Dict, Any, Optional

# third-party will be imported lazily with helpful errors
# ---------------------------
# Config
# ---------------------------
OUTPUT_FILENAME = "log_analysis.xlsx"
BACKUP_TS_FMT = "%Y%m%d_%H%M%S"

# Allowed log names mapping: user can enter number or name (case-insensitive)
LOG_CHOICES = [
    ("Application", "Application"),
    ("Security", "Security"),
    ("Setup", "Setup"),
    ("System", "System"),
    ("ForwardedEvents", "Forwarded Events"),
]


# ---------------------------
# Input parsing / prompts
# ---------------------------

def parse_datetime_input(s: str) -> Optional[Tuple[datetime.datetime, str]]:
    s = s.strip()
    if not s:
        return None
    formats = ["%Y-%m-%d", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"]
    for fmt in formats:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt, fmt
        except ValueError:
            continue
    return None


def prompt_log_choice() -> str:
    print("Choose which Windows Event Log to read from (type number or name). Type 'q' to quit.")
    for i, (internal, friendly) in enumerate(LOG_CHOICES, start=1):
        print(f"  {i}. {friendly}")
    while True:
        choice = input("Enter choice: ").strip()
        if choice.lower() == 'q':
            raise KeyboardInterrupt("User quit")
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(LOG_CHOICES):
                selected = LOG_CHOICES[idx][0]
                print(f"Selected: {LOG_CHOICES[idx][1]}")
                return selected
            else:
                print("Number out of range, try again.")
                continue
        for internal, friendly in LOG_CHOICES:
            if choice.lower() in (internal.lower(), friendly.lower(), friendly.replace(" ", "").lower()):
                print(f"Selected: {friendly}")
                return internal
        print("Unrecognized choice. Enter the number or the log name (e.g. Application).")


def prompt_datetime_range() -> Tuple[datetime.datetime, datetime.datetime]:
    now = datetime.datetime.now()
    print("\nEnter start and end datetimes for the query window.")
    print("Accepted formats: 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM' or 'YYYY-MM-DD HH:MM:SS'.")
    print("Examples: 2024-09-01 00:00   OR   2024-09-01\nType 'q' to quit at any prompt.\n")

    while True:
        raw_start = input("Start datetime: ").strip()
        if raw_start.lower() == 'q':
            raise KeyboardInterrupt("User quit")
        parsed_start = parse_datetime_input(raw_start)
        if not parsed_start:
            print("Could not parse start datetime. Please follow the allowed formats.")
            continue
        start_dt, start_fmt = parsed_start

        raw_end = input("End datetime: ").strip()
        if raw_end.lower() == 'q':
            raise KeyboardInterrupt("User quit")
        parsed_end = parse_datetime_input(raw_end)
        if not parsed_end:
            print("Could not parse end datetime. Please follow the allowed formats.")
            continue
        end_dt, end_fmt = parsed_end

        # If end was provided as date-only, bump to 23:59:59 to cover the whole day.
        if end_fmt == "%Y-%m-%d":
            end_dt = end_dt.replace(hour=23, minute=59, second=59)

        if start_dt > end_dt:
            print("Error: start datetime is after end datetime. Please re-enter.")
            continue

        if end_dt > now:
            print("Error: end datetime is in the future. Please supply a timeframe fully in the past.")
            continue
        if start_dt > now:
            print("Error: start datetime is in the future. Please supply a timeframe fully in the past.")
            continue

        print(f"Accepted range: {start_dt}  ->  {end_dt} (system local time assumed)")
        return start_dt, end_dt


# ---------------------------
# Event-reading helpers
# ---------------------------

def _event_time_to_datetime(ev_time) -> datetime.datetime:
    if isinstance(ev_time, datetime.datetime):
        return ev_time

    fmt_fn = getattr(ev_time, "Format", None)
    if callable(fmt_fn):
        try:
            s = fmt_fn()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
                try:
                    return datetime.datetime.strptime(s, fmt)
                except Exception:
                    continue
        except Exception:
            pass

    try:
        ts = float(ev_time)
        return datetime.datetime.fromtimestamp(ts)
    except Exception:
        pass

    try:
        s = str(ev_time)
        try:
            return datetime.datetime.fromisoformat(s)
        except Exception:
            pass
        for fmt in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.datetime.strptime(s, fmt)
            except Exception:
                continue
    except Exception:
        pass

    raise ValueError(f"Unable to convert event time object to datetime: {repr(ev_time)}")


# basic map for event types -> human string (values fetched from win32evtlog when available)
def _event_type_to_label(event_type, win32evtlog_module=None) -> str:
    """
    Map numeric level values to the requested strings:
      1 -> CRITICAL
      2 -> ERROR
      3 -> WARNING
      4 -> INFORMATIONAL
      5 -> VERBOSE

    If an unknown numeric value is seen, fall back to attempting to use
    win32evtlog constants (if win32evtlog_module provided) or return the value
    as string.
    """
    # Prefer the explicit mapping the user asked for
    explicit_map = {
        1: "CRITICAL",
        2: "ERROR",
        3: "WARNING",
        4: "INFORMATIONAL",
        5: "VERBOSE",
    }
    try:
        # if event_type is e.g. a win32 constant integer, map it directly
        if isinstance(event_type, int):
            if event_type in explicit_map:
                return explicit_map[event_type]
            # if module provided, try mapping common win32evtlog constants to their numeric values
            if win32evtlog_module:
                # attempt to detect common event type constants and map them
                # (this preserves behavior for other numeric codes)
                module_map = {
                    getattr(win32evtlog_module, "EVENTLOG_ERROR_TYPE", None): explicit_map.get(2, "ERROR"),
                    getattr(win32evtlog_module, "EVENTLOG_WARNING_TYPE", None): explicit_map.get(3, "WARNING"),
                    getattr(win32evtlog_module, "EVENTLOG_INFORMATION_TYPE", None): explicit_map.get(4, "INFORMATIONAL"),
                    getattr(win32evtlog_module, "EVENTLOG_AUDIT_SUCCESS", None): "AUDIT_SUCCESS",
                    getattr(win32evtlog_module, "EVENTLOG_AUDIT_FAILURE", None): "AUDIT_FAILURE",
                }
                # module_map keys may include None; handle robustly
                label = module_map.get(event_type)
                if label:
                    return label
            # fallback to string of the numeric value (or unknown)
            return str(event_type)
    except Exception:
        pass

    # If event_type is not int (could be string/other), fall back to previous heuristic
    if win32evtlog_module:
        fallback = {
            getattr(win32evtlog_module, "EVENTLOG_ERROR_TYPE", 1): "ERROR",
            getattr(win32evtlog_module, "EVENTLOG_WARNING_TYPE", 2): "WARNING",
            getattr(win32evtlog_module, "EVENTLOG_INFORMATION_TYPE", 4): "INFORMATION",
            getattr(win32evtlog_module, "EVENTLOG_AUDIT_SUCCESS", 8): "AUDIT_SUCCESS",
            getattr(win32evtlog_module, "EVENTLOG_AUDIT_FAILURE", 16): "AUDIT_FAILURE",
        }
        return fallback.get(event_type, str(event_type))

    # final fallback
    fallback_simple = {
        1: "ERROR",
        2: "WARNING",
        4: "INFORMATION",
        8: "AUDIT_SUCCESS",
        16: "AUDIT_FAILURE",
    }
    return fallback_simple.get(event_type, str(event_type))


def aggregate_events_and_write_excel(log_name: str, start_dt: datetime.datetime, end_dt: datetime.datetime,
                                     out_filename: str = OUTPUT_FILENAME) -> None:
    """
    Read the event log, aggregate by normalized EventID (EventID & 0xFFFF), and
    write an Excel report. Existing file is backed up first.

    After writing via pandas, load with openpyxl and apply UI-friendly formatting:
      - Freeze top header row
      - Bold header, light fill
      - Auto-filter enabled
      - Wrap text for Description column
      - Compute column widths from content (with sensible min/max limits)
      - Set header row height and default row height
      - Set zoomScale to help fit the table on one screen
      - Page setup: fit to 1 page width for printing
    """
    try:
        import win32evtlog
        import win32evtlogutil
    except Exception as e:
        print("ERROR: pywin32 (win32evtlog, win32evtlogutil) is required for aggregation.")
        print("Install with: pip install pywin32")
        raise

    # pandas/openpyxl for writing Excel
    try:
        import pandas as pd
    except Exception:
        print("ERROR: pandas (and openpyxl) are required to write Excel reports.")
        print("Install with: pip install pandas openpyxl")
        raise

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        print("ERROR: openpyxl is required for Excel formatting. Install with: pip install openpyxl")
        raise

    flags = win32evtlog.EVENTLOG_FORWARDS_READ | win32evtlog.EVENTLOG_SEQUENTIAL_READ
    server = None

    handle = None
    scanned = 0
    matched = 0

    aggr: Dict[Any, Dict[str, Any]] = {}

    try:
        handle = win32evtlog.OpenEventLog(server, log_name)
    except Exception as e:
        raise RuntimeError(
            "Failed to open event log '{}': {}\nHint: reading some logs (like 'Security') may require elevated/admin privileges."
            .format(log_name, e)
        ) from e

    try:
        while True:
            events = win32evtlog.ReadEventLog(handle, flags, 0)
            if not events:
                break
            for ev in events:
                scanned += 1
                try:
                    if hasattr(ev, 'TimeGenerated'):
                        evt_time = _event_time_to_datetime(ev.TimeGenerated)
                    elif hasattr(ev, 'TimeWritten'):
                        evt_time = _event_time_to_datetime(ev.TimeWritten)
                    else:
                        continue
                except Exception:
                    continue

                if not (start_dt <= evt_time <= end_dt):
                    continue

                matched += 1

                raw_eid = getattr(ev, "EventID", None)
                try:
                    normalized_eid = (int(raw_eid) & 0xFFFF) if raw_eid is not None else None
                except Exception:
                    normalized_eid = None

                source = getattr(ev, "SourceName", None) or "<Unknown Source>"
                ev_type = getattr(ev, "EventType", None)
                level_label = _event_type_to_label(ev_type, win32evtlog_module=win32evtlog)

                task_cat = None
                if hasattr(ev, "EventCategory"):
                    try:
                        task_cat = getattr(ev, "EventCategory")
                    except Exception:
                        task_cat = None

                description = None
                try:
                    description = win32evtlogutil.SafeFormatMessage(ev, log_name)
                    if description:
                        description = str(description).strip()
                except Exception:
                    inserts = getattr(ev, "StringInserts", None)
                    if inserts:
                        try:
                            description = " ".join([str(x) for x in inserts if x is not None]).strip()
                        except Exception:
                            description = None
                if not description:
                    description = "<No Description Available>"

                key = normalized_eid
                entry = aggr.get(key)
                if entry is None:
                    entry = {
                        "Frequency": 0,
                        "Sources": set(),
                        "Levels": set(),
                        "TaskCategories": set(),
                        "Timestamps": [],
                        "Descriptions": [],
                    }
                    aggr[key] = entry

                entry["Frequency"] += 1
                entry["Sources"].add(source)
                if level_label:
                    entry["Levels"].add(level_label)
                if task_cat is not None:
                    entry["TaskCategories"].add(str(task_cat))
                entry["Timestamps"].append(evt_time)
                entry["Descriptions"].append(description)
    finally:
        if handle:
            try:
                win32evtlog.CloseEventLog(handle)
            except Exception:
                pass

    # Build rows for DataFrame
    rows = []
    si = 1
    for eventid, data in sorted(aggr.items(), key=lambda kv: (kv[0] if kv[0] is not None else -1, -kv[1]["Frequency"])):
        sources_str = " || ".join(sorted(data["Sources"]))
        levels_str = " || ".join(sorted(data["Levels"]))
        taskcat_str = " || ".join(sorted(data["TaskCategories"])) if data["TaskCategories"] else ""
        ts_str = ", ".join([t.isoformat(" ") for t in data["Timestamps"]])
        desc_str = "\n----------\n".join(data["Descriptions"])
        rows.append({
            "SI no": si,
            "EventID": eventid,
            "Source": sources_str,
            "Level": levels_str,
            "Task Category": taskcat_str,
            "Timestamp (logged)": ts_str,
            "Description": desc_str,
            "Frequency": data["Frequency"],
        })
        si += 1

    df = pd.DataFrame(rows, columns=["SI no", "EventID", "Source", "Level", "Task Category", "Timestamp (logged)", "Description", "Frequency"])

    # Backup existing Excel if present
    if os.path.exists(out_filename):
        ts = datetime.datetime.now().strftime(BACKUP_TS_FMT)
        backup_name = f"{os.path.splitext(out_filename)[0]}_{ts}{os.path.splitext(out_filename)[1]}"
        try:
            shutil.move(out_filename, backup_name)
            print(f"Backed up existing '{out_filename}' -> '{backup_name}'")
        except Exception as e:
            print(f"Warning: failed to back up existing '{out_filename}': {e}")

    # Write Excel using pandas (openpyxl engine)
    try:
        df.to_excel(out_filename, index=False)
        print(f"Wrote report to '{out_filename}' ({len(df)} unique EventIDs).")
    except Exception as e:
        print(f"ERROR: failed to write Excel file '{out_filename}': {e}")
        raise

    # ============================================
    # Post-process workbook with openpyxl formatting
    # ============================================
    try:
        wb = openpyxl.load_workbook(out_filename)
        ws = wb.active  # first sheet

        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # Apply header styles
        for col_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Freeze header row
        ws.freeze_panes = ws["A2"]

        # Auto-filter for the header row
        ws.auto_filter.ref = ws.dimensions

        # Default row height and header row height
        ws.row_dimensions[1].height = 26
        default_row_height = 16
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = default_row_height

        # Compute column widths based on max length of content (with limits)
        # We'll treat Description and Timestamp specially
        max_col_width = 80  # upper cap (reasonable for most screens)
        min_col_width = 8

        for idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(idx)
            # Gather max length of column header and cell contents
            max_len = len(str(col))
            for cell in ws[col_letter]:
                if cell.value is None:
                    continue
                # For multi-line cells, consider the longest line
                s = str(cell.value)
                # replace newlines to estimate width of the longest line
                longest_line = max((len(line) for line in s.splitlines()), default=len(s))
                if longest_line > max_len:
                    max_len = longest_line
            # Special adjustments
            if col == "Description":
                width = min(max(max_len + 4, 30), max_col_width)  # Description should be wide but capped
                # Wrap text for description column
                for cell in ws[col_letter]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col == "Timestamp (logged)":
                width = min(max(max_len + 4, 25), max_col_width)
                for cell in ws[col_letter]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                width = min(max(max_len + 2, min_col_width), max_col_width)
                for cell in ws[col_letter]:
                    cell.alignment = Alignment(wrap_text=False, vertical="center")
            ws.column_dimensions[col_letter].width = width

        # Set sheet view zoom to fit more columns on a single screen (user can still change in Excel)
        try:
            ws.sheet_view.zoomScale = 80  # percent
        except Exception:
            # some openpyxl versions differ; ignore if fails
            pass

        # Page setup: fit to 1 page wide (useful when printing)
        try:
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            pass

        # Save workbook
        wb.save(out_filename)
        print(f"Applied formatting to '{out_filename}' (freeze header, filters, column widths, wrap description).")
    except Exception as e:
        print(f"Warning: failed to apply Excel formatting: {e}")
        # do not raise — the raw file is still present

    # Print summary
    print("\n--- Aggregation Summary ---")
    print(f"Total events scanned during aggregation: {scanned}")
    print(f"Events matched in window               : {matched}")
    print(f"Unique EventIDs found                  : {len(df)}")
    print("-----------------------------\n")

# ---------------------------
# Main flow
# ---------------------------

def main():
    print("=== Windows Event Log Analyzer ===\n")
    if platform.system() != 'Windows':
        print("ERROR: This script must be run on Windows to read Event Logs (pywin32).")
        print("You can still test prompts elsewhere but reading/aggregation will fail.")
    try:
        selected_log = prompt_log_choice()
        start_dt, end_dt = prompt_datetime_range()
    except KeyboardInterrupt:
        print("\nOperation cancelled by user. Exiting.")
        sys.exit(0)
    except Exception:
        print("\nAn unexpected error occurred while collecting inputs:")
        traceback.print_exc()
        sys.exit(1)

    # First quick read to count scans and matches (keeps small memory) -- reuse earlier logic
    try:
        # We'll reuse the read loop but not keep heavy data; use aggregate function directly because it reads and aggregates in one pass.
        print("\nProceed to aggregate and write Excel report now? (y/n)")
        ans = input("> ").strip().lower()
        if ans not in ("y", "yes"):
            print("Skipping aggregation. Exiting after input step.")
            sys.exit(0)

        aggregate_events_and_write_excel(selected_log, start_dt, end_dt, out_filename=OUTPUT_FILENAME)
    except RuntimeError as re:
        print("\nRuntime error while attempting to access event logs:")
        print(re)
        print("\nIf you are on Windows and haven't installed pywin32, run: pip install pywin32")
        print("If the log is 'Security' you may need to run this script as Administrator.")
        sys.exit(1)
    except Exception:
        print("\nUnexpected error during aggregation:")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
